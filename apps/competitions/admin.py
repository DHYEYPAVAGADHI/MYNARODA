"""
Competition Registrations — Enhanced Admin
==========================================
Approval workflow, status badges, per-type tabs, premium Excel export.
"""

import datetime
from django.contrib import admin, messages
from django.http import HttpResponse
from django.utils.html import format_html
from django.utils.translation import gettext_lazy as _
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from unfold.admin import ModelAdmin
from unfold.contrib.filters.admin import ChoicesDropdownFilter
from unfold.decorators import action

from .models import CompetitionRegistration
from .utils import generate_organization_certificate
import zipfile
import io


@admin.register(CompetitionRegistration)
class CompetitionRegistrationAdmin(ModelAdmin):
    list_display = (
        "registration_id",
        "organization_name",
        "org_type_badge",
        "city",
        "authorized_person_name",
        "mobile_number",
        "status_badge",
        "solar_badge",
        "rainwater_badge",
        "certificate_status",
        "created_at",
    )
    list_filter = (
        ("organization_type", ChoicesDropdownFilter),
        ("status", ChoicesDropdownFilter),
        "city",
        "certificate_generated",
        "solar_installed",
        "rainwater_harvesting",
    )
    search_fields = (
        "registration_id",
        "organization_name",
        "authorized_person_name",
        "email",
        "mobile_number",
        "city",
    )
    readonly_fields = ("registration_id", "created_at", "updated_at")
    list_per_page = 25
    date_hierarchy = "created_at"
    ordering = ("-created_at",)
    list_filter_submit = True
    actions = [
        "approve_registrations",
        "hold_registrations",
        "reject_registrations",
        "export_excel_all",
        "export_excel_schools",
        "export_excel_colleges",
        "export_excel_societies",
        "export_excel_ngos",
        "regenerate_certificates",
        "download_certificates_zip",
    ]

    fieldsets = (
        (_("Registration Info"), {
            "fields": ("registration_id", "status", "created_at", "updated_at"),
        }),
        (_("Certificate"), {
            "fields": ("certificate_generated", "certificate_png", "certificate_pdf"),
        }),
        (_("Organization Details"), {
            "fields": (
                "organization_type", "organization_name", "registration_number",
                "address", "city", "pincode",
            ),
        }),
        (_("Residential Society — Additional Info"), {
            "fields": ("homes_count", "members_participating"),
            "classes": ("collapse",),
        }),
        (_("Authorized Contact Person"), {
            "fields": (
                "authorized_person_name", "designation",
                "mobile_number", "whatsapp_number", "alternate_number", "email",
            ),
        }),
        (_("Sustainability"), {
            "fields": ("solar_installed", "solar_panels_count", "rainwater_harvesting"),
        }),
        (_("Supporting Evidence"), {
            "fields": ("supporting_file",),
        }),
    )

    # ── Display ────────────────────────────────────────────────────

    def status_badge(self, obj):
        colors = {
            "APPROVED": ("#DCFCE7", "#16A34A", "✓"),
            "PENDING": ("#FFF7ED", "#EA580C", "⏳"),
            "REJECTED": ("#FEF2F2", "#DC2626", "✗"),
        }
        bg, fg, icon = colors.get(obj.status, ("#F3F4F6", "#6B7280", "?"))
        return format_html(
            '<span style="background:{};color:{};padding:3px 10px;border-radius:999px;'
            'font-size:0.7rem;font-weight:700;">{} {}</span>',
            bg, fg, icon, obj.get_status_display()
        )
    status_badge.short_description = "Status"

    def org_type_badge(self, obj):
        if not obj.organization_type:
            return format_html('<span style="color:#6B7280;">Unknown</span>')
        name = obj.organization_type.name
        icon, bg, fg = ("🏢", "#F9FAFB", "#6B7280")
        if "school" in name.lower():
            icon, bg, fg = ("🏫", "#EFF6FF", "#2563EB")
        elif "college" in name.lower():
            icon, bg, fg = ("🎓", "#F5F3FF", "#7C3AED")
        elif "society" in name.lower():
            icon, bg, fg = ("🏘️", "#F0FDFA", "#0D9488")
        elif "ngo" in name.lower() or "trust" in name.lower():
            icon, bg, fg = ("🤝", "#FFF7ED", "#EA580C")
        
        return format_html(
            '<span style="background:{};color:{};padding:3px 10px;border-radius:999px;'
            'font-size:0.7rem;font-weight:700;">{} {}</span>',
            bg, fg, icon, name
        )
    org_type_badge.short_description = "Type"

    def solar_badge(self, obj):
        if obj.solar_installed:
            return format_html('<span title="Solar Installed" style="color:#F97316;font-size:1rem;">☀️</span>')
        return format_html('<span style="color:#E5E7EB;">☀️</span>')
    solar_badge.short_description = "Solar"

    @admin.display(description="Rainwater")
    def rainwater_badge(self, obj):
        if obj.rainwater_harvesting:
            return format_html('<span style="color:#0EA5E9;">💧 Yes</span>')
        return format_html('<span style="color:#9CA3AF;">-</span>')

    @admin.display(description="Certificate")
    def certificate_status(self, obj):
        if obj.certificate_generated:
            return format_html('<span style="background:#DCFCE7; color:#16A34A; padding:2px 8px; border-radius:12px; font-size:11px; font-weight:600;">Generated</span>')
        return format_html('<span style="background:#F3F4F6; color:#6B7280; padding:2px 8px; border-radius:12px; font-size:11px; font-weight:600;">Pending</span>')

    # ── Actions ────────────────────────────────────────────────────

    @action(description=_("✅ Approve selected registrations"))
    def approve_registrations(self, request, queryset):
        count = queryset.update(status=CompetitionRegistration.ApprovalStatus.APPROVED)
        messages.success(request, _(f"{count} registrations approved."))

    @action(description=_("⏸️ Put selected registrations On Hold"))
    def hold_registrations(self, request, queryset):
        # Hold maps to PENDING status (no new model needed)
        count = queryset.update(status=CompetitionRegistration.ApprovalStatus.PENDING)
        messages.info(request, _(f"{count} registrations set to Pending/Hold."))

    @action(description="Reject selected registrations")
    def reject_registrations(self, request, queryset):
        updated = queryset.update(status=CompetitionRegistration.ApprovalStatus.REJECTED)
        self.message_user(request, f"{updated} registrations rejected.", level=messages.ERROR)

    @action(description="Regenerate Certificates (PNG & PDF)")
    def regenerate_certificates(self, request, queryset):
        count = 0
        for reg in queryset:
            generate_organization_certificate(reg)
            count += 1
        self.message_user(request, f"Successfully regenerated {count} certificates.", level=messages.SUCCESS)

    @action(description="Download ZIP of Selected Certificates")
    def download_certificates_zip(self, request, queryset):
        buffer = io.BytesIO()
        with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for reg in queryset:
                if reg.certificate_pdf:
                    zip_file.write(reg.certificate_pdf.path, f"Certificates/{reg.registration_id}.pdf")
                if reg.certificate_png:
                    zip_file.write(reg.certificate_png.path, f"Certificates/{reg.registration_id}.png")
        
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/zip')
        response['Content-Disposition'] = f'attachment; filename="certificates_{datetime.date.today()}.zip"'
        return response

    @action(description=_("📊 Export ALL selected to Excel"))
    def export_excel_all(self, request, queryset):
        return _export_competition_excel(queryset, label="all")

    @action(description=_("🏫 Export SCHOOLS to Excel"))
    def export_excel_schools(self, request, queryset):
        qs = queryset.filter(organization_type__name__icontains="School")
        return _export_competition_excel(qs, label="schools")

    @action(description=_("🎓 Export COLLEGES to Excel"))
    def export_excel_colleges(self, request, queryset):
        qs = queryset.filter(organization_type__name__icontains="College")
        return _export_competition_excel(qs, label="colleges")

    @action(description=_("🏘️ Export SOCIETIES to Excel"))
    def export_excel_societies(self, request, queryset):
        qs = queryset.filter(organization_type__name__icontains="Society")
        return _export_competition_excel(qs, label="societies")

    @action(description=_("🤝 Export NGOs to Excel"))
    def export_excel_ngos(self, request, queryset):
        qs = queryset.filter(organization_type__name__icontains="NGO")
        return _export_competition_excel(qs, label="ngos")


def _export_competition_excel(queryset, label="registrations"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = label.title()
    ws.freeze_panes = "A2"

    header_fill = PatternFill(start_color="0B7A3B", end_color="0B7A3B", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = [
        "Registration ID", "Status", "Organization Type", "Organization Name",
        "Registration Number", "Address", "City", "Pincode",
        "Homes Count", "Members Participating",
        "Contact Person", "Designation", "Mobile", "WhatsApp",
        "Alternate Number", "Email",
        "Solar Installed", "Solar Panels", "Rainwater Harvesting",
        "Registration Date",
    ]

    ws.append(headers)
    ws.row_dimensions[1].height = 26

    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    for reg in queryset:
        org_type_name = reg.organization_type.name if reg.organization_type else "Unknown"
        ws.append([
            reg.registration_id,
            reg.get_status_display(),
            org_type_name,
            reg.organization_name,
            reg.registration_number or "",
            reg.address,
            reg.city,
            reg.pincode,
            reg.homes_count or "",
            reg.members_participating or "",
            reg.authorized_person_name,
            reg.designation or "",
            reg.mobile_number,
            reg.whatsapp_number,
            reg.alternate_number or "",
            reg.email,
            "Yes" if reg.solar_installed else "No",
            reg.solar_panels_count or "",
            "Yes" if reg.rainwater_harvesting else "No",
            reg.created_at.strftime("%Y-%m-%d %H:%M") if reg.created_at else "",
        ])

    for col_idx, col in enumerate(ws.columns, 1):
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 42)

    today = datetime.date.today().strftime("%Y-%m-%d")
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{label}_{today}.xlsx"'
    wb.save(response)
    return response

