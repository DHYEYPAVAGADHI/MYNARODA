"""
Campaign Admin Panel — Views
============================
Custom Mission Control admin panel. All views restricted to ADMIN/SUPER_ADMIN.
Provides: login, dashboard, pledges, students, organizations, gallery,
events, news, analytics, settings, and Excel exports.
"""
from __future__ import annotations

import csv
import io
import json
import base64
from datetime import timedelta

from django.contrib import messages
from django.core.files.base import ContentFile
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.paginator import Paginator
from django.db.models import Count, Q
from django.db.models.functions import TruncDate, TruncMonth
from django.http import Http404, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.utils.translation import gettext_lazy as _
from django.views import View

from apps.accounts.models import User, UserRole
from apps.competitions.models import CompetitionRegistration
from apps.events.models import Event
from apps.gallery.models import Photo
from apps.cms.models import LeadershipPhotos, HarGharTirangaRegistration
from apps.news.models import NewsArticle

from apps.student_portal.models import StudentSubmission
from apps.volunteers.models import PledgeRegistration


try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _is_admin(user) -> bool:
    return (
        user.is_authenticated
        and (
            user.is_superuser
            or getattr(user, "role", None) in (UserRole.ADMIN, UserRole.SUPER_ADMIN)
        )
    )


def _xl_response(filename: str):
    """Return an HttpResponse configured for .xlsx download."""
    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


def _style_xl_header(ws, headers: list[str], fill_color: str = "0B5D2A"):
    """Write and style the header row of an Excel sheet."""
    header_fill = PatternFill("solid", fgColor=fill_color)
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(header) + 6, 16)
    ws.row_dimensions[1].height = 22


# ─── Mixin ────────────────────────────────────────────────────────────────────

class AdminRequiredMixin:
    """Restrict view to ADMIN and SUPER_ADMIN roles only."""

    def dispatch(self, request, *args, **kwargs):
        if not _is_admin(request.user):
            return redirect("admin_panel:login")
        return super().dispatch(request, *args, **kwargs)


# ─── Login / Logout ───────────────────────────────────────────────────────────

class AdminLoginView(View):
    template_name = "admin_panel/login.html"

    def get(self, request):
        if _is_admin(request.user):
            return redirect("admin_panel:dashboard")
        return render(request, self.template_name)

    def post(self, request):
        email = request.POST.get("email", "").strip().lower()
        password = request.POST.get("password", "")

        user = authenticate(request, email=email, password=password)
        if user is not None and _is_admin(user):
            login(request, user)
            return redirect(request.GET.get("next", "admin_panel:dashboard"))

        return render(
            request,
            self.template_name,
            {
                "email": email,
                "login_error": "Invalid email or password. Please try again.",
                "shake_form": True
            }
        )

def axes_lockout_view(request, credentials=None):
    """Custom view for django-axes when a user is locked out."""
    email = request.POST.get("email", "").strip().lower()
    return render(
        request,
        "admin_panel/login.html",
        {
            "email": email,
            "lockout_error": "Too many failed attempts. Please wait 30 seconds and try again.",
            "shake_form": True
        },
        status=429
    )


class AdminLogoutView(View):
    def post(self, request):
        logout(request)
        return redirect("admin_panel:login")


# ─── Dashboard ────────────────────────────────────────────────────────────────

class AdminDashboardView(AdminRequiredMixin, View):
    template_name = "admin_panel/dashboard.html"

    def get(self, request):
        now = timezone.now()

        # 30-day pledge trend
        trend_days, trend_counts = [], []
        for i in range(29, -1, -1):
            d = now - timedelta(days=i)
            trend_days.append(d.strftime("%b %d"))
            trend_counts.append(
                PledgeRegistration.objects.filter(created_at__date=d.date()).count()
            )

        # Ward participation (via ward field if exists, else org name)
        ward_qs = (
            PledgeRegistration.objects
            .values("organization__name")
            .annotate(cnt=Count("id"))
            .order_by("-cnt")[:8]
        )

        # Recent pledges
        recent_pledges = PledgeRegistration.objects.select_related("organization").order_by("-created_at")[:6]

        # Upcoming events
        upcoming_events = Event.objects.filter(
            starts_at__gte=now, is_published=True
        ).order_by("starts_at")[:4]

        context = {
            "total_pledges": PledgeRegistration.objects.count(),
            "total_students": StudentSubmission.objects.count(),
            "total_orgs": CompetitionRegistration.objects.count(),
            "approved_orgs": CompetitionRegistration.objects.filter(status="APPROVED").count(),
            "total_photos": Photo.objects.count(),
            "pending_photos": Photo.objects.filter(approval_status="PENDING").count(),
            "total_events": Event.objects.count(),
            "total_tiranga": HarGharTirangaRegistration.objects.count(),

            # Charts
            "chart_days": json.dumps(trend_days),
            "chart_counts": json.dumps(trend_counts),
            "ward_labels": json.dumps([w["organization__name"] or "—" for w in ward_qs]),
            "ward_counts": json.dumps([w["cnt"] for w in ward_qs]),

            "recent_pledges": recent_pledges,
            "upcoming_events": upcoming_events,
        }
        return render(request, self.template_name, context)


# ─── Pledges ──────────────────────────────────────────────────────────────────

class AdminPledgeListView(AdminRequiredMixin, View):
    template_name = "admin_panel/pledges.html"

    def get(self, request):
        qs = PledgeRegistration.objects.select_related("organization").order_by("-created_at")

        q = request.GET.get("q", "").strip()
        org = request.GET.get("org", "").strip()
        status = request.GET.get("status", "").strip()
        date_from = request.GET.get("date_from", "").strip()
        date_to = request.GET.get("date_to", "").strip()

        if q:
            qs = qs.filter(Q(full_name__icontains=q) | Q(mobile_number__icontains=q))
        if org:
            qs = qs.filter(organization__name__icontains=org)
        if status == "approved":
            qs = qs.filter(is_approved=True)
        elif status == "pending":
            qs = qs.filter(is_approved=False)
        if date_from:
            qs = qs.filter(created_at__date__gte=date_from)
        if date_to:
            qs = qs.filter(created_at__date__lte=date_to)

        paginator = Paginator(qs, 25)
        page_obj = paginator.get_page(request.GET.get("page", 1))

        return render(request, self.template_name, {
            "page_obj": page_obj,
            "total": qs.count(),
            "q": q, "org": org, "status": status,
            "date_from": date_from, "date_to": date_to,
        })


# ─── Students ─────────────────────────────────────────────────────────────────

class AdminStudentListView(AdminRequiredMixin, View):
    template_name = "admin_panel/students.html"

    def get(self, request):
        qs = StudentSubmission.objects.order_by("-created_at")

        q = request.GET.get("q", "").strip()
        school = request.GET.get("school", "").strip()
        comp_type = request.GET.get("type", "").strip()
        status = request.GET.get("status", "").strip()

        if q:
            qs = qs.filter(
                Q(student_name__icontains=q) | Q(guardian_mobile__icontains=q) | Q(school_name__icontains=q)
            )
        if school:
            qs = qs.filter(school_name__icontains=school)
        if comp_type:
            qs = qs.filter(competition_type=comp_type)
        if status:
            qs = qs.filter(status=status)

        paginator = Paginator(qs, 25)
        page_obj = paginator.get_page(request.GET.get("page", 1))

        return render(request, self.template_name, {
            "page_obj": page_obj,
            "total": qs.count(),
            "total_approved": StudentSubmission.objects.filter(status="APPROVED").count(),
            "q": q, "school": school, "comp_type": comp_type, "status": status,
            "type_choices": StudentSubmission.CompetitionType.choices,
            "status_choices": StudentSubmission.StatusChoices.choices,
        })


class AdminStudentActionView(AdminRequiredMixin, View):
    """Approve or reject a student submission."""

    def post(self, request, pk):
        submission = get_object_or_404(StudentSubmission, pk=pk)
        action = request.POST.get("action")
        if action == "approve":
            submission.status = StudentSubmission.StatusChoices.APPROVED
            submission.save(update_fields=["status"])
            messages.success(request, f"Submission by {submission.student_name} approved.")
        elif action == "reject":
            submission.status = StudentSubmission.StatusChoices.REJECTED
            submission.save(update_fields=["status"])
            messages.success(request, f"Submission by {submission.student_name} rejected.")
        return redirect("admin_panel:students")


# ─── Organizations ────────────────────────────────────────────────────────────

class AdminOrgListView(AdminRequiredMixin, View):
    template_name = "admin_panel/organizations.html"

    def get(self, request):
        qs = CompetitionRegistration.objects.order_by("-created_at")

        q = request.GET.get("q", "").strip()
        status = request.GET.get("status", "").strip()
        org_type = request.GET.get("type", "").strip()

        if q:
            qs = qs.filter(
                Q(organization_name__icontains=q) | Q(authorized_person_name__icontains=q)
            )
        if status:
            qs = qs.filter(status=status)
        if org_type:
            qs = qs.filter(organization_type_id=org_type)

        paginator = Paginator(qs, 18)
        page_obj = paginator.get_page(request.GET.get("page", 1))

        # Dynamically build choices from CompetitionOrganizationType
        from apps.competitions.models import CompetitionOrganizationType
        try:
            org_types = CompetitionOrganizationType.objects.filter(is_active=True).order_by('sort_order', 'name')
            type_choices = [(str(t.id), t.name) for t in org_types]
        except Exception:
            type_choices = []

        try:
            total_count = CompetitionRegistration.objects.count()
            approved_count = CompetitionRegistration.objects.filter(status="APPROVED").count()
            pending_count = CompetitionRegistration.objects.filter(status="PENDING").count()
            status_choices = CompetitionRegistration.ApprovalStatus.choices
        except Exception:
            total_count = 0
            approved_count = 0
            pending_count = 0
            status_choices = []

        return render(request, self.template_name, {
            "page_obj": page_obj,
            "total": total_count,
            "approved": approved_count,
            "pending": pending_count,
            "q": q, "status": status, "org_type": org_type,
            "type_choices": type_choices,
            "status_choices": status_choices,
        })


class AdminOrgActionView(AdminRequiredMixin, View):
    """Approve, reject, or toggle org status."""

    def post(self, request, pk):
        org = get_object_or_404(CompetitionRegistration, pk=pk)
        action = request.POST.get("action")
        if action == "approve":
            org.status = CompetitionRegistration.ApprovalStatus.APPROVED
            org.save(update_fields=["status"])
            messages.success(request, f"'{org.organization_name}' approved.")
        elif action == "reject":
            org.status = CompetitionRegistration.ApprovalStatus.REJECTED
            org.save(update_fields=["status"])
            messages.success(request, f"'{org.organization_name}' rejected.")
        return redirect("admin_panel:organizations")


# ─── Gallery / Photo Queue ────────────────────────────────────────────────────

class AdminGalleryView(AdminRequiredMixin, View):
    template_name = "admin_panel/gallery.html"

    def get(self, request):
        status_filter = request.GET.get("status", "PENDING")
        qs = Photo.objects.filter(approval_status=status_filter).select_related(
            "photographer", "category"
        ).order_by("-created_at")

        paginator = Paginator(qs, 24)
        page_obj = paginator.get_page(request.GET.get("page", 1))

        return render(request, self.template_name, {
            "page_obj": page_obj,
            "status_filter": status_filter,
            "pending_count": Photo.objects.filter(approval_status="PENDING").count(),
            "approved_count": Photo.objects.filter(approval_status="APPROVED").count(),
            "rejected_count": Photo.objects.filter(approval_status="REJECTED").count(),
        })

    def post(self, request):
        action = request.POST.get("action")
        photo_ids = request.POST.getlist("photo_ids")
        if not photo_ids:
            messages.warning(request, "No photos selected.")
            return redirect("admin_panel:gallery")

        photos = Photo.objects.filter(id__in=photo_ids)
        if action == "approve":
            photos.update(
                approval_status=Photo.ApprovalStatus.APPROVED,
                approved_by=request.user,
                approved_at=timezone.now(),
            )
            messages.success(request, f"{photos.count()} photo(s) approved.")
        elif action == "reject":
            photos.update(approval_status=Photo.ApprovalStatus.REJECTED)
            messages.success(request, f"{photos.count()} photo(s) rejected.")
        elif action == "feature":
            photos.update(is_featured=True)
            messages.success(request, f"{photos.count()} photo(s) featured.")

        return redirect("admin_panel:gallery")


class AdminPhotoActionView(AdminRequiredMixin, View):
    """Single-photo approve/reject/feature via POST."""

    def post(self, request, photo_id):
        photo = get_object_or_404(Photo, id=photo_id)
        action = request.POST.get("action")
        if action == "approve":
            photo.approval_status = Photo.ApprovalStatus.APPROVED
            photo.approved_by = request.user
            photo.approved_at = timezone.now()
            photo.save()
            messages.success(request, "Photo approved.")
        elif action == "reject":
            photo.approval_status = Photo.ApprovalStatus.REJECTED
            photo.save()
            messages.success(request, "Photo rejected.")
        elif action == "feature":
            photo.is_featured = not photo.is_featured
            photo.save()
            messages.success(request, f"Photo {'featured' if photo.is_featured else 'unfeatured'}.")
        return redirect("admin_panel:gallery")


# ─── Events ───────────────────────────────────────────────────────────────────

class AdminEventsView(AdminRequiredMixin, View):
    template_name = "admin_panel/events.html"

    def get(self, request):
        qs = Event.objects.order_by("-starts_at")
        q = request.GET.get("q", "").strip()
        status = request.GET.get("status", "").strip()
        if q:
            qs = qs.filter(Q(title__icontains=q) | Q(location_name__icontains=q))
        if status:
            qs = qs.filter(status=status)

        paginator = Paginator(qs, 20)
        page_obj = paginator.get_page(request.GET.get("page", 1))

        return render(request, self.template_name, {
            "page_obj": page_obj,
            "total": Event.objects.count(),
            "upcoming": Event.objects.filter(status="UPCOMING").count(),
            "q": q, "status": status,
            "status_choices": Event.EventStatus.choices,
        })


class AdminEventTogglePublishView(AdminRequiredMixin, View):
    def post(self, request, pk):
        event = get_object_or_404(Event, pk=pk)
        event.is_published = not event.is_published
        event.save(update_fields=["is_published"])
        state = "published" if event.is_published else "unpublished"
        messages.success(request, f"Event '{event.title}' {state}.")
        return redirect("admin_panel:events")



# ─── Analytics (JSON API) ─────────────────────────────────────────────────────

class AdminAnalyticsView(AdminRequiredMixin, View):
    template_name = "admin_panel/analytics.html"

    def get(self, request):
        now = timezone.now()

        # 6-month pledge trend
        months, pledge_counts, student_counts, org_counts = [], [], [], []
        for i in range(5, -1, -1):
            d = now - timedelta(days=30 * i)
            label = d.strftime("%b %Y")
            months.append(label)
            pledge_counts.append(
                PledgeRegistration.objects.filter(created_at__year=d.year, created_at__month=d.month).count()
            )
            student_counts.append(
                StudentSubmission.objects.filter(created_at__year=d.year, created_at__month=d.month).count()
            )
            org_counts.append(
                CompetitionRegistration.objects.filter(created_at__year=d.year, created_at__month=d.month).count()
            )

        # Ward / org participation (top 10)
        ward_qs = (
            PledgeRegistration.objects
            .values("organization__name")
            .annotate(cnt=Count("id"))
            .order_by("-cnt")[:10]
        )

        context = {
            "months_json": json.dumps(months),
            "pledge_counts_json": json.dumps(pledge_counts),
            "student_counts_json": json.dumps(student_counts),
            "org_counts_json": json.dumps(org_counts),
            "ward_labels_json": json.dumps([w["organization__name"] or "—" for w in ward_qs]),
            "ward_counts_json": json.dumps([w["cnt"] for w in ward_qs]),

            # Summary numbers
            "total_pledges": PledgeRegistration.objects.count(),
            "total_students": StudentSubmission.objects.count(),
            "total_orgs": CompetitionRegistration.objects.count(),
            "total_photos": Photo.objects.count(),
        }
        return render(request, self.template_name, context)


# ─── Settings ─────────────────────────────────────────────────────────────────

class AdminSettingsView(AdminRequiredMixin, View):
    template_name = "admin_panel/settings.html"

    def get(self, request):
        return render(request, self.template_name, {
            "admins": User.objects.filter(
                role__in=[UserRole.ADMIN, UserRole.SUPER_ADMIN]
            ).order_by("-created_at"),
        })


# ─── Excel Exports ────────────────────────────────────────────────────────────

class AdminExportPledgesExcelView(AdminRequiredMixin, View):
    def get(self, request):
        qs = PledgeRegistration.objects.select_related("organization").order_by("-created_at")

        if not HAS_OPENPYXL:
            return HttpResponse("openpyxl not installed.", status=500)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Pledges"

        headers = [
            "Sr.", "Certificate ID", "Full Name", "Mobile", "Email",
            "Gender", "Age", "City", "Organization", "Approved", "Date",
        ]
        _style_xl_header(ws, headers)

        for i, p in enumerate(qs, 1):
            ws.append([
                i,
                p.certificate_id or "",
                p.full_name,
                p.mobile_number,
                p.email,
                p.get_gender_display(),
                p.age or "",
                p.city,
                p.organization.name if p.organization else "",
                "Yes" if p.is_approved else "No",
                p.created_at.strftime("%Y-%m-%d %H:%M"),
            ])

        resp = _xl_response("pledge_registrations.xlsx")
        wb.save(resp)
        return resp


class AdminExportStudentsExcelView(AdminRequiredMixin, View):
    def get(self, request):
        qs = StudentSubmission.objects.order_by("-created_at")

        if not HAS_OPENPYXL:
            return HttpResponse("openpyxl not installed.", status=500)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Students"

        headers = [
            "Sr.", "Participation ID", "Student Name", "Parent Name",
            "Mobile", "School", "Grade", "Competition Type", "Status", "Date",
        ]
        _style_xl_header(ws, headers)

        for i, s in enumerate(qs, 1):
            ws.append([
                i,
                s.participation_id,
                s.student_name,
                s.parent_name,
                s.guardian_mobile,
                s.school_name,
                s.grade,
                s.get_competition_type_display(),
                s.get_status_display(),
                s.created_at.strftime("%Y-%m-%d"),
            ])

        resp = _xl_response("student_registrations.xlsx")
        wb.save(resp)
        return resp


class AdminExportOrgsExcelView(AdminRequiredMixin, View):
    def get(self, request):
        qs = CompetitionRegistration.objects.order_by("-created_at")

        if not HAS_OPENPYXL:
            return HttpResponse("openpyxl not installed.", status=500)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Organizations"

        headers = [
            "Sr.", "Registration ID", "Organization Name", "Type",
            "Authorized Person", "Designation", "Mobile", "Email",
            "City", "Status", "Date",
        ]
        _style_xl_header(ws, headers)

        for i, o in enumerate(qs, 1):
            ws.append([
                i,
                o.registration_id if hasattr(o, "registration_id") else str(o.pk),
                o.organization_name,
                o.get_organization_type_display(),
                o.authorized_person_name,
                o.designation,
                o.mobile_number,
                o.email,
                o.city,
                o.get_status_display(),
                o.created_at.strftime("%Y-%m-%d"),
            ])

        resp = _xl_response("organization_registrations.xlsx")
        wb.save(resp)
        return resp


class AdminExportPledgesCSVView(AdminRequiredMixin, View):
    """CSV export for pledges (legacy / backup)."""
    def get(self, request):
        qs = PledgeRegistration.objects.select_related("organization").order_by("-created_at")
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="pledge_registrations.csv"'
        writer = csv.writer(response)
        writer.writerow([
            "Certificate ID", "Full Name", "Mobile", "Email",
            "Gender", "Age", "City", "Organization", "Approved", "Date",
        ])
        for p in qs:
            writer.writerow([
                p.certificate_id or "",
                p.full_name,
                p.mobile_number,
                p.email,
                p.get_gender_display(),
                p.age or "",
                p.city,
                p.organization.name if p.organization else "",
                "Yes" if p.is_approved else "No",
                p.created_at.strftime("%Y-%m-%d %H:%M"),
            ])
        return response


# ─── Public Stats API ─────────────────────────────────────────────────────────

class PublicStatsAPIView(View):
    """JSON endpoint for public homepage live counters. No auth required."""

    def get(self, request):
        return JsonResponse({
            "pledges": PledgeRegistration.objects.count(),
            "students": StudentSubmission.objects.count(),
            "organizations": CompetitionRegistration.objects.filter(status="APPROVED").count(),
            "photos": Photo.objects.filter(approval_status="APPROVED").count(),
            "events": Event.objects.filter(is_published=True).count(),
        })


class AdminLeadershipPhotosView(AdminRequiredMixin, View):
    template_name = "admin_panel/leadership_photos.html"

    def get(self, request):
        photos = LeadershipPhotos.objects.first()
        
        committee_leaders = [
            {"name": "Dr. Payalben Kukrani", "role": "President", "field": "president", "model_field": "payalben_photo"},
            {"name": "Shri Nikunj Rameshbhai Khakhi", "role": "Convener", "field": "convener", "model_field": "nikunj_photo"},
        ]

        co_convener_leaders = [
            {"name": "Shri Kiran Raval", "role": "President, BJP Naroda Ward", "field": "kiran_raval", "model_field": "kiran_raval_photo"},
            {"name": "Shri Vipul Patel", "role": "Chairman & Municipal Councilor, Naroda AMC", "field": "vipul_patel", "model_field": "vipul_patel_photo"},
            {"name": "Shri Jayeshbhai Prajapati", "role": "Municipal Councilor, Naroda AMC", "field": "jayesh_prajapati", "model_field": "jayesh_prajapati_photo"},
            {"name": "Shri Chandaben Patel", "role": "Municipal Councilor, Naroda AMC", "field": "chandaben_patel", "model_field": "chandaben_patel_photo"},
            {"name": "Shri Divya Nikunj Khakhi", "role": "Dy. Chairman & Municipal Councilor, Naroda AMC", "field": "divya_khakhi", "model_field": "divya_khakhi_photo"},
        ]


        leaders = [
            {"name": "Shri Nitin Nabin", "role": "Hon'ble National President, BJP", "field": "nitin_nabin_photo", "model_field": "nitin_nabin_photo"},
            {"name": "Shri Bhupendrabhai Patel", "role": "Hon'ble Chief Minister, Gujarat", "field": "bhupendrabhai_patel_photo", "model_field": "bhupendrabhai_patel_photo"},
            {"name": "Shri Jagdish Vishwakarma", "role": "Hon'ble State President, BJP Gujarat", "field": "jagdish_vishwakarma_photo", "model_field": "jagdish_vishwakarma_photo"},
            {"name": "Shri Harshbhai Sanghavi", "role": "Hon'ble Dy. Chief Minister of Gujarat", "field": "harshbhai_sanghavi_photo", "model_field": "harshbhai_sanghavi_photo"},
            {"name": "Shri Ratnakarji", "role": "Hon'ble State Organization General Secretary", "field": "ratnakarji_photo", "model_field": "ratnakarji_photo"},
            {"name": "Shri Ajay Brahmbhatt", "role": "State General Secretary, BJP Gujarat", "field": "ajay_brahmbhatt_photo", "model_field": "ajay_brahmbhatt_photo"},
            {"name": "Shri Anirudhbhai Dave", "role": "State General Secretary, BJP Gujarat", "field": "anirudhbhai_dave_photo", "model_field": "anirudhbhai_dave_photo"},
            {"name": "Dr. Prashantbhai Korat", "role": "State General Secretary, BJP Gujarat", "field": "prashantbhai_korat_photo", "model_field": "prashantbhai_korat_photo"},
            {"name": "Shri Hitendrasinh Chauhan", "role": "State General Secretary, BJP Gujarat", "field": "hitendrasinh_chauhan_photo", "model_field": "hitendrasinh_chauhan_photo"},
            {"name": "Shri Prerakbhai Shah", "role": "President, Karnavati Mahanagar BJP", "field": "prerakbhai_shah_photo", "model_field": "prerakbhai_shah_photo"},
            {"name": "Shri Hasmukhbhai Patel", "role": "Hon. Member of Parliament", "field": "hasmukhbhai_patel_photo", "model_field": "hasmukhbhai_patel_photo"},
            {"name": "Shri Dineshbhai Makwana", "role": "Hon. Member of Parliament", "field": "dineshbhai_makwana_photo", "model_field": "dineshbhai_makwana_photo"},
        ]
        
        # Attach the photo URL if it exists
        if photos:
            for leader in leaders + committee_leaders + co_convener_leaders:
                model_attr = leader.get("model_field", leader["field"])
                img_field = getattr(photos, model_attr, None)
                if img_field and img_field.name:
                    leader["photo"] = img_field

        return render(request, self.template_name, {"leaders": leaders, "committee_leaders": committee_leaders, "co_convener_leaders": co_convener_leaders})

    def post(self, request):
        from django.http import JsonResponse
        is_ajax = request.POST.get('ajax_upload') == '1'
        
        photos = LeadershipPhotos.objects.first()
        if not photos:
            photos = LeadershipPhotos.objects.create()

        field_map = {
            'president': 'payalben_photo',
            'convener': 'nikunj_photo',
            'kiran_raval': 'kiran_raval_photo',
            'vipul_patel': 'vipul_patel_photo',
            'jayesh_prajapati': 'jayesh_prajapati_photo',
            'chandaben_patel': 'chandaben_patel_photo',
            'divya_khakhi': 'divya_khakhi_photo',
            'nitin_nabin_photo': 'nitin_nabin_photo',
            'bhupendrabhai_patel_photo': 'bhupendrabhai_patel_photo',
            'jagdish_vishwakarma_photo': 'jagdish_vishwakarma_photo',
            'harshbhai_sanghavi_photo': 'harshbhai_sanghavi_photo',
            'ratnakarji_photo': 'ratnakarji_photo',
            'ajay_brahmbhatt_photo': 'ajay_brahmbhatt_photo',
            'anirudhbhai_dave_photo': 'anirudhbhai_dave_photo',
            'prashantbhai_korat_photo': 'prashantbhai_korat_photo',
            'hitendrasinh_chauhan_photo': 'hitendrasinh_chauhan_photo',
            'prerakbhai_shah_photo': 'prerakbhai_shah_photo',
            'hasmukhbhai_patel_photo': 'hasmukhbhai_patel_photo',
            'dineshbhai_makwana_photo': 'dineshbhai_makwana_photo'
        }
        
        updated = False
        
        # Check for standard cropped blob upload
        uploaded = request.FILES.get('photo')
        field_key = request.POST.get('field_key')
        
        if field_key in field_map and uploaded:
            setattr(photos, field_map[field_key], uploaded)
            photos.save()
            if is_ajax:
                url = getattr(photos, field_map[field_key]).url if getattr(photos, field_map[field_key]) else ''
                return JsonResponse({"success": True, "url": url, "message": "Photo uploaded successfully."})
            
            messages.success(request, "Photo uploaded successfully.")
            return redirect('admin_panel:leadership_photos')
            
        # Handle remove logic
        remove_field = request.POST.get("remove_field")
        if remove_field:
            if remove_field in field_map:
                setattr(photos, field_map[remove_field], None)
                updated = True
            elif remove_field in field_map.values():
                setattr(photos, remove_field, None)
                updated = True
                
            if is_ajax and updated:
                photos.save()
                return JsonResponse({"success": True, "message": "Photo removed."})
        
        # Handle regular file uploads (fallback)
        for key, model_field in field_map.items():
            if key in request.FILES:
                setattr(photos, model_field, request.FILES[key])
                updated = True
            elif model_field in request.FILES:
                setattr(photos, model_field, request.FILES[model_field])
                updated = True
                
        if updated:
            photos.save()
            if is_ajax:
                return JsonResponse({"success": True, "message": "Leadership photo updated successfully."})
            messages.success(request, "Leadership photo updated successfully.")

        if is_ajax:
            return JsonResponse({"success": False, "error": "No valid action found."})

        return redirect("admin_panel:leadership_photos")


# ─── News ─────────────────────────────────────────────────────────────────────

class AdminNewsView(AdminRequiredMixin, View):
    template_name = "admin_panel/news.html"

    def get(self, request):
        qs = NewsArticle.objects.select_related("category").order_by("-published_at", "-id")
        q = request.GET.get("q", "").strip()
        status = request.GET.get("status", "").strip()
        if q:
            qs = qs.filter(Q(title__icontains=q) | Q(summary__icontains=q))
        if status == "published":
            qs = qs.filter(is_published=True)
        elif status == "draft":
            qs = qs.filter(is_published=False)

        paginator = Paginator(qs, 20)
        page_obj = paginator.get_page(request.GET.get("page", 1))

        return render(request, self.template_name, {
            "page_obj": page_obj,
            "total": NewsArticle.objects.count(),
            "published": NewsArticle.objects.filter(is_published=True).count(),
            "q": q, "status": status,
        })


class AdminNewsTogglePublishView(AdminRequiredMixin, View):
    def post(self, request, pk):
        article = get_object_or_404(NewsArticle, pk=pk)
        article.is_published = not article.is_published
        if article.is_published and not article.published_at:
            article.published_at = timezone.now()
        article.save(update_fields=["is_published", "published_at"])
        state = "published" if article.is_published else "unpublished"
        messages.success(request, f"Article '{article.title}' {state}.")
        return redirect("admin_panel:news")


# ─── Branding Settings ────────────────────────────────────────────────────────

from apps.cms.models import BrandingSettings

class AdminBrandingHeaderLogosView(AdminRequiredMixin, View):
    template_name = "admin_panel/branding_header_logos.html"

    def get(self, request):
        settings_obj = BrandingSettings.load()
        
        # Determine fallback URLs based on the template
        fallbacks = {
            "my_naroda_logo": "/static/images/logo-mynaroda.jpeg",
            "bjp_logo": "/static/images/logo-bjp.png",
            "pratham_logo": "/static/images/logo-pratham.png",
            "gncn_logo": "/static/images/logo-gncn.png",
        }
        
        return render(request, self.template_name, {
            "branding_settings": settings_obj,
            "fallbacks": fallbacks
        })

    def post(self, request):
        settings_obj = BrandingSettings.load()
        
        try:
            data = json.loads(request.body)
            action = data.get("action")
            field = data.get("field")
            
            valid_fields = ["my_naroda_logo", "bjp_logo", "pratham_logo", "gncn_logo"]
            if field not in valid_fields:
                return JsonResponse({"success": False, "error": "Invalid field"})
            
            if action == "save":
                image_data = data.get("image_data")
                if not image_data:
                    return JsonResponse({"success": False, "error": "No image data provided"})
                
                # Format: data:image/png;base64,iVBORw0KGgo...
                format, imgstr = image_data.split(';base64,') 
                ext = format.split('/')[-1] 
                
                # Validate extension
                if ext not in ['png', 'jpg', 'jpeg', 'webp']:
                    ext = 'png'
                
                data_file = ContentFile(base64.b64decode(imgstr), name=f"{field}.{ext}")
                
                # Delete old if exists
                old_file = getattr(settings_obj, field)
                if old_file:
                    old_file.delete(save=False)
                
                setattr(settings_obj, field, data_file)
                settings_obj.save()
                
                # Get the new URL
                new_url = getattr(settings_obj, field).url
                
                return JsonResponse({"success": True, "url": new_url})
                
            elif action == "remove":
                # Delete file
                old_file = getattr(settings_obj, field)
                if old_file:
                    old_file.delete(save=False)
                
                setattr(settings_obj, field, None)
                settings_obj.save()
                return JsonResponse({"success": True})
                
            return JsonResponse({"success": False, "error": "Unknown action"})
            
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})


# ─── Har Ghar Tiranga ─────────────────────────────────────────────────────────
from apps.cms.models import HarGharTirangaRegistration

class AdminHarGharTirangaRegistrationsView(AdminRequiredMixin, View):
    template_name = "admin_panel/tiranga_registrations.html"

    def get(self, request):
        qs = HarGharTirangaRegistration.objects.all().order_by("-created_at")
        
        # simple search
        q = request.GET.get("q", "").strip()
        if q:
            qs = qs.filter(
                Q(token_id__icontains=q) |
                Q(name__icontains=q) |
                Q(mobile_number__icontains=q) |
                Q(address__icontains=q)
            )

        paginator = Paginator(qs, 50)
        page_obj = paginator.get_page(request.GET.get("page", 1))
        
        today = timezone.now().date()
        today_count = HarGharTirangaRegistration.objects.filter(created_at__date=today).count()

        return render(request, self.template_name, {
            "page_obj": page_obj,
            "total": qs.count(),
            "today_count": today_count,
            "q": q,
        })


class AdminExportHarGharTirangaExcelView(AdminRequiredMixin, View):
    def get(self, request):
        qs = HarGharTirangaRegistration.objects.all().order_by("-created_at")

        if not HAS_OPENPYXL:
            return HttpResponse("openpyxl not installed.", status=500)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Tiranga Registrations"

        headers = [
            "Sr.", "Token ID", "Name", "Mobile Number",
            "Address", "Location", "Registration Date",
        ]
        _style_xl_header(ws, headers)

        for i, reg in enumerate(qs, 1):
            location_str = "Not provided"
            if reg.latitude and reg.longitude:
                location_str = f"{reg.latitude}, {reg.longitude}"
                if reg.location_text:
                    location_str += f" ({reg.location_text})"
                    
            ws.append([
                i,
                reg.token_id or "-",
                reg.name,
                reg.mobile_number,
                reg.address,
                location_str,
                reg.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            ])

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="har_ghar_tiranga_registrations.xlsx"'
        wb.save(response)
        return response


# ─── Organization Types & Associate Organizations ─────────────────────────────
from apps.competitions.models import CompetitionOrganizationType
from apps.volunteers.models import Organization
import json

class AdminOrganizationTypesView(AdminRequiredMixin, View):
    template_name = "admin_panel/org_types.html"

    def get(self, request):
        qs = CompetitionOrganizationType.objects.all().order_by("sort_order", "name")
        q = request.GET.get("q", "").strip()
        if q:
            qs = qs.filter(name__icontains=q)
        return render(request, self.template_name, {"items": qs, "q": q})

    def post(self, request):
        action = request.POST.get("action")
        
        if action == "add":
            name = request.POST.get("name", "").strip()
            is_active = request.POST.get("is_active") == "on"
            sort_order = request.POST.get("sort_order", 0)
            if not name:
                messages.error(request, "Name cannot be empty.")
            elif CompetitionOrganizationType.objects.filter(name__iexact=name).exists():
                messages.error(request, "Organization type with this name already exists.")
            else:
                try:
                    sort_order = int(sort_order)
                    CompetitionOrganizationType.objects.create(name=name, is_active=is_active, sort_order=sort_order)
                    messages.success(request, "Organization type added.")
                except ValueError:
                    messages.error(request, "Sort order must be numeric.")

        elif action == "edit":
            pk = request.POST.get("id")
            name = request.POST.get("name", "").strip()
            is_active = request.POST.get("is_active") == "on"
            sort_order = request.POST.get("sort_order", 0)
            
            try:
                obj = CompetitionOrganizationType.objects.get(pk=pk)
                if not name:
                    messages.error(request, "Name cannot be empty.")
                elif CompetitionOrganizationType.objects.filter(name__iexact=name).exclude(pk=pk).exists():
                    messages.error(request, "Organization type with this name already exists.")
                else:
                    obj.name = name
                    obj.is_active = is_active
                    obj.sort_order = int(sort_order)
                    obj.save()
                    messages.success(request, "Organization type updated.")
            except Exception as e:
                messages.error(request, f"Error updating: {e}")

        elif action == "delete":
            pk = request.POST.get("id")
            CompetitionOrganizationType.objects.filter(pk=pk).delete()
            messages.success(request, "Organization type deleted.")
            
        elif action == "toggle":
            pk = request.POST.get("id")
            obj = CompetitionOrganizationType.objects.filter(pk=pk).first()
            if obj:
                obj.is_active = not obj.is_active
                obj.save()
                messages.success(request, f"Organization type {'enabled' if obj.is_active else 'disabled'}.")

        return redirect("admin_panel:org_types")


class AdminAssociateOrganizationsView(AdminRequiredMixin, View):
    template_name = "admin_panel/associate_orgs.html"

    def get(self, request):
        qs = Organization.objects.all().order_by("sort_order", "name")
        q = request.GET.get("q", "").strip()
        if q:
            qs = qs.filter(name__icontains=q)
        return render(request, self.template_name, {"items": qs, "q": q})

    def post(self, request):
        action = request.POST.get("action")
        
        if action == "add":
            name = request.POST.get("name", "").strip()
            is_active = request.POST.get("is_active") == "on"
            sort_order = request.POST.get("sort_order", 0)
            if not name:
                messages.error(request, "Name cannot be empty.")
            elif Organization.objects.filter(name__iexact=name).exists():
                messages.error(request, "Organization with this name already exists.")
            else:
                try:
                    sort_order = int(sort_order)
                    Organization.objects.create(name=name, is_active=is_active, sort_order=sort_order)
                    messages.success(request, "Associate organization added.")
                except ValueError:
                    messages.error(request, "Sort order must be numeric.")

        elif action == "edit":
            pk = request.POST.get("id")
            name = request.POST.get("name", "").strip()
            is_active = request.POST.get("is_active") == "on"
            sort_order = request.POST.get("sort_order", 0)
            
            try:
                obj = Organization.objects.get(pk=pk)
                if not name:
                    messages.error(request, "Name cannot be empty.")
                elif Organization.objects.filter(name__iexact=name).exclude(pk=pk).exists():
                    messages.error(request, "Organization with this name already exists.")
                else:
                    obj.name = name
                    obj.is_active = is_active
                    obj.sort_order = int(sort_order)
                    obj.save()
                    messages.success(request, "Associate organization updated.")
            except Exception as e:
                messages.error(request, f"Error updating: {e}")

        elif action == "delete":
            pk = request.POST.get("id")
            Organization.objects.filter(pk=pk).delete()
            messages.success(request, "Associate organization deleted.")
            
        elif action == "toggle":
            pk = request.POST.get("id")
            obj = Organization.objects.filter(pk=pk).first()
            if obj:
                obj.is_active = not obj.is_active
                obj.save()
                messages.success(request, f"Associate organization {'enabled' if obj.is_active else 'disabled'}.")

        return redirect("admin_panel:associate_orgs")


# ─── Bulk Uploads ─────────────────────────────────────────────────────────────

import uuid
from datetime import datetime, date
from django.utils.crypto import get_random_string

class AdminTirangaSampleExcelView(AdminRequiredMixin, View):
    def get(self, request):
        if not HAS_OPENPYXL:
            messages.error(request, "Excel export/import requires openpyxl.")
            return redirect("admin_panel:tiranga_registrations")
            
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sample Registrations"
        
        headers = ["Full Name", "Mobile", "Address", "Location", "Registration Date"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="0E7A37", end_color="0E7A37", fill_type="solid")
            
        ws.append(["Rahul Desai", "9876543210", "123 Naroda Road", "Naroda Patiya", "2024-08-01"])
        
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = 'attachment; filename="tiranga_sample.xlsx"'
        wb.save(response)
        return response

class AdminTirangaUploadExcelView(AdminRequiredMixin, View):
    def post(self, request):
        if not HAS_OPENPYXL:
            messages.error(request, "Excel export/import requires openpyxl.")
            return redirect("admin_panel:tiranga_registrations")
            
        file = request.FILES.get("excel_file")
        if not file or not file.name.endswith('.xlsx'):
            messages.error(request, "Please upload a valid .xlsx file.")
            return redirect("admin_panel:tiranga_registrations")
            
        if file.size > 5 * 1024 * 1024:
            messages.error(request, "File size must be under 5MB.")
            return redirect("admin_panel:tiranga_registrations")
            
        try:
            wb = openpyxl.load_workbook(file, data_only=True)
            ws = wb.active
            
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                messages.error(request, "File is empty.")
                return redirect("admin_panel:tiranga_registrations")
                
            headers = [str(h).strip().lower() if h else "" for h in rows[0]]
            
            col_map = {}
            for i, h in enumerate(headers):
                if 'token id' in h: col_map['token'] = i
                elif 'citizen details' in h: col_map['citizen'] = i
                elif 'full name' in h or h == 'name': col_map['name'] = i
                elif 'mobile' in h or 'phone' in h: col_map['mobile'] = i
                elif 'address' in h: col_map['address'] = i
                elif 'location' in h: col_map['location'] = i
                
            is_format_a = 'citizen' in col_map
            is_format_b = 'name' in col_map and 'mobile' in col_map
            
            if not is_format_a and not is_format_b:
                messages.error(request, "Unrecognized column format. Please make sure the header row contains valid columns.")
                return redirect("admin_panel:tiranga_registrations")
                
            imported = 0
            skipped = 0
            skip_reasons = []
            
            for row in rows[1:]:
                if not any(row):
                    continue
                    
                def get_val(key):
                    idx = col_map.get(key)
                    if idx is not None and idx < len(row) and row[idx] is not None:
                        return str(row[idx]).strip()
                    return ""
                    
                token_id = get_val('token')
                address = get_val('address')
                location = get_val('location')
                
                name, mobile = "", ""
                if is_format_a:
                    citizen_details = get_val('citizen')
                    if '|' in citizen_details:
                        parts = citizen_details.split('|')
                        name = parts[0].strip()
                        mobile = parts[1].strip()
                    else:
                        name = citizen_details.strip()
                        mobile = ""
                else:
                    name = get_val('name')
                    mobile = get_val('mobile')
                
                if not name or not mobile:
                    skipped += 1
                    skip_reasons.append("Missing name or mobile")
                    continue
                    
                # Clean mobile number
                mobile_clean = "".join(filter(str.isdigit, mobile))
                if len(mobile_clean) != 10:
                    skipped += 1
                    skip_reasons.append(f"Invalid mobile number: {name} ({mobile})")
                    continue
                
                if HarGharTirangaRegistration.objects.filter(name=name, mobile_number=mobile_clean).exists():
                    skipped += 1
                    skip_reasons.append(f"Duplicate entry: {name} ({mobile_clean})")
                    continue
                
                if not token_id:
                    token_id = f"TIRANGA-80-{get_random_string(6).upper()}"
                    while HarGharTirangaRegistration.objects.filter(token_id=token_id).exists():
                        token_id = f"TIRANGA-80-{get_random_string(6).upper()}"
                        
                HarGharTirangaRegistration.objects.create(
                    token_id=token_id,
                    name=name,
                    mobile_number=mobile_clean,
                    address=address,
                    location_text=location
                )
                imported += 1
                
            msg = f"Imported {imported} registrations. Skipped {skipped} rows."
            if skip_reasons:
                msg += f" Reasons: {', '.join(skip_reasons[:3])}..."
                messages.warning(request, msg)
            else:
                messages.success(request, msg)
                
        except Exception as e:
            messages.error(request, f"Error processing file: {str(e)}")
            
        return redirect("admin_panel:tiranga_registrations")


from apps.volunteers.models import Organization, OrganizationType

class AdminPledgeSampleExcelView(AdminRequiredMixin, View):
    def get(self, request):
        if not HAS_OPENPYXL:
            messages.error(request, "Excel export/import requires openpyxl.")
            return redirect("admin_panel:pledges")
            
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sample Pledges"
        
        headers = ["Sr.", "Full Name", "Mobile", "Email", "Gender", "Organization", "City", "Status", "Date"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="0E7A37", end_color="0E7A37", fill_type="solid")
            
        ws.append([1, "Rahul Desai", "9876543210", "rahul@example.com", "MALE", "Lions Club", "Ahmedabad", "Approved", "2024-08-01"])
        
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = 'attachment; filename="pledges_sample.xlsx"'
        wb.save(response)
        return response

class AdminPledgeUploadExcelView(AdminRequiredMixin, View):
    def post(self, request):
        if not HAS_OPENPYXL:
            messages.error(request, "Excel export/import requires openpyxl.")
            return redirect("admin_panel:pledges")
            
        file = request.FILES.get("excel_file")
        if not file or not file.name.endswith('.xlsx'):
            messages.error(request, "Please upload a valid .xlsx file.")
            return redirect("admin_panel:pledges")
            
        if file.size > 5 * 1024 * 1024:
            messages.error(request, "File size must be under 5MB.")
            return redirect("admin_panel:pledges")
            
        try:
            wb = openpyxl.load_workbook(file, data_only=True)
            ws = wb.active
            
            imported = 0
            skipped = 0
            skip_reasons = []
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                    
                sr, name, mobile, email, gender, org_name, city, status, reg_date = row[:9]
                
                if not name or not mobile:
                    skipped += 1
                    skip_reasons.append("Missing name or mobile")
                    continue
                    
                mobile = str(mobile).strip()
                name = str(name).strip()
                
                if not mobile.isdigit() or len(mobile) != 10:
                    skipped += 1
                    skip_reasons.append(f"Invalid mobile number: {mobile}")
                    continue
                
                if PledgeRegistration.objects.filter(full_name=name, mobile_number=mobile).exists():
                    skipped += 1
                    skip_reasons.append(f"Duplicate pledge: {name} ({mobile})")
                    continue
                
                org_obj = None
                if org_name:
                    org_obj, _ = Organization.objects.get_or_create(
                        name=str(org_name).strip(),
                        defaults={"org_type": OrganizationType.OTHERS}
                    )
                else:
                    org_obj, _ = Organization.objects.get_or_create(
                        name="Individual / General Citizen",
                        defaults={"org_type": OrganizationType.OTHERS}
                    )
                        
                is_approved = str(status).strip().lower() == "approved" if status else True
                
                # Check gender
                gender = str(gender).upper().strip() if gender else "OTHER"
                if gender not in [c[0] for c in PledgeRegistration.GenderChoices.choices]:
                    gender = "OTHER"
                
                # Create token/cert ID equivalent
                cert_id = f"GNCN{get_random_string(6).upper()}"
                while PledgeRegistration.objects.filter(certificate_id=cert_id).exists():
                    cert_id = f"GNCN{get_random_string(6).upper()}"
                
                PledgeRegistration.objects.create(
                    full_name=name,
                    mobile_number=mobile,
                    email=str(email) if email else "",
                    gender=gender,
                    organization=org_obj,
                    city=str(city) if city else "Ahmedabad",
                    is_approved=is_approved,
                    date_of_birth=date(2000, 1, 1),
                    certificate_id=cert_id,
                    agreed_to_pledge=True,
                    consent_accepted=True,
                    otp_verified=True
                )
                imported += 1
                
            msg = f"Imported {imported} pledges. Skipped {skipped} rows."
            if skip_reasons:
                msg += f" Reasons: {', '.join(skip_reasons[:3])}..."
                messages.warning(request, msg)
            else:
                messages.success(request, msg)
                
        except Exception as e:
            messages.error(request, f"Error processing file: {str(e)}")

        return redirect("admin_panel:pledges")


# ─── Hero Banner ────────────────────────────────────────────────────────────

class AdminHeroBannerView(AdminRequiredMixin, View):
    """Manage the homepage hero banner slides (title/subtitle/description,
    image, buttons, active state, and display order)."""
    template_name = "admin_panel/hero_banner.html"

    def get(self, request):
        from apps.cms.models import Homepage
        homepage = Homepage.load()
        slides = homepage.hero_slides.all().order_by("sort_order", "-created_at")
        return render(request, self.template_name, {"slides": slides})

    def post(self, request):
        from apps.cms.models import Homepage, HeroSlide, MediaAsset

        action = request.POST.get("action")
        homepage = Homepage.load()

        if action == "create":
            image_asset = None
            image_file = request.FILES.get("image")
            if image_file:
                image_asset = MediaAsset.objects.create(
                    category=MediaAsset.AssetCategory.HERO,
                    title=request.POST.get("title") or image_file.name,
                    image=image_file,
                )
            HeroSlide.objects.create(
                homepage=homepage,
                title=request.POST.get("title", "").strip(),
                subtitle=request.POST.get("subtitle", "").strip(),
                description=request.POST.get("description", "").strip(),
                image=image_asset,
                primary_button_text=request.POST.get("primary_button_text", "").strip(),
                primary_button_url=request.POST.get("primary_button_url", "").strip(),
                secondary_button_text=request.POST.get("secondary_button_text", "").strip(),
                secondary_button_url=request.POST.get("secondary_button_url", "").strip(),
                is_active=request.POST.get("is_active") == "on",
                sort_order=request.POST.get("sort_order") or 0,
            )
            messages.success(request, "Hero slide added.")

        elif action == "update":
            slide = get_object_or_404(HeroSlide, pk=request.POST.get("slide_id"), homepage=homepage)
            slide.title = request.POST.get("title", "").strip()
            slide.subtitle = request.POST.get("subtitle", "").strip()
            slide.description = request.POST.get("description", "").strip()
            slide.primary_button_text = request.POST.get("primary_button_text", "").strip()
            slide.primary_button_url = request.POST.get("primary_button_url", "").strip()
            slide.secondary_button_text = request.POST.get("secondary_button_text", "").strip()
            slide.secondary_button_url = request.POST.get("secondary_button_url", "").strip()
            slide.is_active = request.POST.get("is_active") == "on"
            slide.sort_order = request.POST.get("sort_order") or 0
            image_file = request.FILES.get("image")
            if image_file:
                slide.image = MediaAsset.objects.create(
                    category=MediaAsset.AssetCategory.HERO,
                    title=slide.title or image_file.name,
                    image=image_file,
                )
            slide.save()
            messages.success(request, "Hero slide updated.")

        elif action == "delete":
            slide = get_object_or_404(HeroSlide, pk=request.POST.get("slide_id"), homepage=homepage)
            slide.delete()
            messages.success(request, "Hero slide deleted.")

        elif action == "toggle":
            slide = get_object_or_404(HeroSlide, pk=request.POST.get("slide_id"), homepage=homepage)
            slide.is_active = not slide.is_active
            slide.save(update_fields=["is_active"])

        return redirect("admin_panel:hero_banner")
