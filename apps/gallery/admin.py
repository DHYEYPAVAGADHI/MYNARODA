"""
Gallery App — Enhanced Admin Configuration
==========================================
Photo thumbnail previews, one-click approve/reject, Excel export.
"""

import datetime
from django.contrib import admin
from django.utils import timezone
from django.utils.html import format_html
from django.utils.translation import gettext_lazy as _
from django.http import HttpResponse
from unfold.admin import ModelAdmin
from unfold.decorators import action
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from apps.gallery.models import GalleryCategory, Photo, PhotoTag, GalleryCollection


@admin.register(GalleryCategory)
class GalleryCategoryAdmin(ModelAdmin):
    list_display = ("name", "slug", "sort_order")
    search_fields = ("name", "description")
    ordering = ("sort_order",)
    prepopulated_fields = {"slug": ("name",)}


@admin.register(PhotoTag)
class PhotoTagAdmin(ModelAdmin):
    list_display = ("name", "slug", "created_at")
    search_fields = ("name",)
    ordering = ("name",)
    prepopulated_fields = {"slug": ("name",)}


@admin.register(Photo)
class PhotoAdmin(ModelAdmin):
    list_display = (
        "photo_preview",
        "title",
        "photographer",
        "category",
        "media_type",
        "status_badge",
        "featured_badge",
        "view_count",
        "created_at",
    )
    list_filter = (
        "approval_status",
        "is_featured",
        "category",
        "media_type",
        ("created_at", admin.DateFieldListFilter),
    )
    search_fields = ("title", "photographer__email", "photographer__full_name", "location")
    filter_horizontal = ("tags",)
    readonly_fields = (
        "created_at", "updated_at", "deleted_at",
        "photo_preview_large", "view_count",
    )
    ordering = ("-created_at",)
    date_hierarchy = "created_at"
    list_per_page = 25
    actions = [
        "approve_photos",
        "reject_photos",
        "feature_photos",
        "unfeature_photos",
        "export_approved_excel",
    ]

    fieldsets = (
        ("Media", {
            "fields": ("photo_preview_large", "cloudinary_id", "url", "local_image", "media_type"),
        }),
        ("Information", {
            "fields": ("title", "location", "taken_at"),
        }),
        ("Classification", {
            "fields": ("category", "tags", "photographer", "event"),
        }),
        ("Approval & Display", {
            "fields": ("approval_status", "is_featured", "approved_by", "approved_at"),
        }),
        ("Analytics", {
            "fields": ("view_count",),
        }),
        ("Audit", {
            "fields": ("created_at", "updated_at", "deleted_at"),
            "classes": ("collapse",),
        }),
    )

    # ── Display ────────────────────────────────────────────────────

    def photo_preview(self, obj):
        url = obj.thumbnail_url or obj.url
        if url:
            return format_html(
                '<img src="{}" style="height:52px;width:72px;object-fit:cover;'
                'border-radius:8px;border:1px solid #E5E7EB;" />',
                url
            )
        if obj.local_image:
            return format_html(
                '<img src="{}" style="height:52px;width:72px;object-fit:cover;'
                'border-radius:8px;border:1px solid #E5E7EB;" />',
                obj.local_image.url
            )
        return format_html('<span style="color:#9CA3AF;font-size:1.5rem;">🖼️</span>')
    photo_preview.short_description = "Preview"

    def photo_preview_large(self, obj):
        url = obj.url or (obj.local_image.url if obj.local_image else None)
        if url:
            return format_html(
                '<img src="{}" style="max-height:300px;max-width:100%;border-radius:12px;'
                'box-shadow:0 4px 20px rgba(0,0,0,0.12);" />',
                url
            )
        return "—"
    photo_preview_large.short_description = "Photo Preview"

    def status_badge(self, obj):
        mapping = {
            "APPROVED": ("#DCFCE7", "#16A34A", "✓ Approved"),
            "PENDING": ("#FFF7ED", "#EA580C", "⏳ Pending"),
            "REJECTED": ("#FEF2F2", "#DC2626", "✗ Rejected"),
        }
        bg, fg, label = mapping.get(obj.approval_status, ("#F3F4F6", "#6B7280", obj.approval_status))
        return format_html(
            '<span style="background:{};color:{};padding:3px 10px;border-radius:999px;'
            'font-size:0.7rem;font-weight:700;">{}</span>',
            bg, fg, label
        )
    status_badge.short_description = "Status"

    def featured_badge(self, obj):
        if obj.is_featured:
            return format_html('<span style="color:#F97316;font-size:1rem;" title="Featured">⭐</span>')
        return format_html('<span style="color:#E5E7EB;font-size:1rem;">⭐</span>')
    featured_badge.short_description = "★"

    # ── Actions ───────────────────────────────────────────────────

    @action(description=_("✅ Approve selected photos"))
    def approve_photos(self, request, queryset):
        count = queryset.update(
            approval_status=Photo.ApprovalStatus.APPROVED,
            approved_by=request.user,
            approved_at=timezone.now(),
        )
        self.message_user(request, f"{count} photos approved and are now live on the website.")

    @action(description=_("❌ Reject selected photos"))
    def reject_photos(self, request, queryset):
        count = queryset.update(approval_status=Photo.ApprovalStatus.REJECTED)
        self.message_user(request, f"{count} photos rejected.")

    @action(description=_("⭐ Feature selected on homepage"))
    def feature_photos(self, request, queryset):
        count = queryset.update(is_featured=True)
        self.message_user(request, f"{count} photos marked as featured.")

    @action(description=_("☆ Remove from featured"))
    def unfeature_photos(self, request, queryset):
        count = queryset.update(is_featured=False)
        self.message_user(request, f"{count} photos removed from featured.")

    @action(description=_("📊 Export approved photos to Excel"))
    def export_approved_excel(self, request, queryset):
        approved_qs = queryset.filter(approval_status=Photo.ApprovalStatus.APPROVED)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Approved Gallery"
        ws.freeze_panes = "A2"

        header_fill = PatternFill(start_color="0B7A3B", end_color="0B7A3B", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        headers = ["Title", "Photographer", "Category", "Location", "Media Type", "View Count", "Approved On", "URL"]
        ws.append(headers)
        for col_num, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font

        for photo in approved_qs:
            ws.append([
                photo.title or "",
                photo.photographer.full_name if photo.photographer else "",
                photo.category.name if photo.category else "",
                photo.location or "",
                photo.get_media_type_display(),
                photo.view_count,
                photo.approved_at.strftime("%Y-%m-%d") if photo.approved_at else "",
                photo.url or "",
            ])

        for col_idx, col in enumerate(ws.columns, 1):
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 50)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        today = datetime.date.today().strftime("%Y-%m-%d")
        response["Content-Disposition"] = f'attachment; filename="gallery_approved_{today}.xlsx"'
        wb.save(response)
        return response


@admin.register(GalleryCollection)
class GalleryCollectionAdmin(ModelAdmin):
    list_display = ("title", "slug", "is_published", "created_at")
    list_filter = ("is_published",)
    search_fields = ("title", "description")
    prepopulated_fields = {"slug": ("title",)}
    filter_horizontal = ("photos",)
    ordering = ("-created_at",)
