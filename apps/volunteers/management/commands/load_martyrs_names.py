"""
Management command to load freedom fighter / martyrs names from an Excel file
into the FreedomFighterName model used for Tree Pledge Certificate assignment.

Usage:
    python manage.py load_martyrs_names /path/to/All_Martyrs_Names_1.xlsx
"""

from django.core.management.base import BaseCommand, CommandError


class Command(BaseCommand):
    help = "Load freedom fighter / martyrs names from an Excel file into the database."

    def add_arguments(self, parser):
        parser.add_argument("excel_path", type=str, help="Absolute path to the .xlsx file")
        parser.add_argument(
            "--clear",
            action="store_true",
            help="Clear all existing names before loading (use with caution)",
        )

    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError:
            raise CommandError("openpyxl is required. Run: pip install openpyxl")

        from apps.volunteers.models import FreedomFighterName

        excel_path = options["excel_path"]

        try:
            wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
            ws = wb.active
        except Exception as e:
            raise CommandError(f"Could not open Excel file: {e}")

        if options["clear"]:
            count = FreedomFighterName.objects.count()
            FreedomFighterName.objects.all().delete()
            self.stdout.write(self.style.WARNING(f"Cleared {count} existing names."))

        # Detect name column (header row)
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        name_col_idx = None
        for i, h in enumerate(headers):
            if h and str(h).strip().lower() == "name":
                name_col_idx = i
                break

        if name_col_idx is None:
            raise CommandError(f"Could not find a 'Name' column in the Excel file. Found columns: {headers}")

        imported = 0
        skipped = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) <= name_col_idx:
                continue
            raw_name = row[name_col_idx]
            if not raw_name:
                continue
            name = str(raw_name).strip()
            if not name:
                continue

            _, created = FreedomFighterName.objects.get_or_create(
                name=name,
                defaults={"is_active": True, "used_in_current_cycle": False},
            )
            if created:
                imported += 1
            else:
                skipped += 1

        self.stdout.write(
            self.style.SUCCESS(
                f"Done! Imported: {imported} new names | Skipped (already exist): {skipped} | "
                f"Total in DB: {FreedomFighterName.objects.count()}"
            )
        )
