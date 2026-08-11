"""
Volunteers App — Enhanced Admin Configuration
==============================================
Premium admin for Pledge Registrations with Excel export,
bulk actions, and rich list display.
"""

import datetime
import io
from django.contrib import admin, messages
from django.http import HttpResponse
from django.utils.html import format_html
from django.utils.translation import gettext_lazy as _
from unfold.admin import ModelAdmin
from unfold.decorators import action
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .models import PledgeRegistration, Organization, FreedomFighterName


@admin.register(Organization)
class OrganizationAdmin(ModelAdmin):
    list_display = ("name", "org_type", "is_active", "created_at")
    list_filter = ("org_type", "is_active")
    list_editable = ("is_active",)
    search_fields = ("name",)
    ordering = ("name",)


@admin.register(FreedomFighterName)
class FreedomFighterNameAdmin(ModelAdmin):
    list_display = ("name", "is_active", "used_in_current_cycle")
    list_filter = ("is_active", "used_in_current_cycle")
    list_editable = ("is_active",)
    search_fields = ("name",)
    ordering = ("name",)


@admin.register(PledgeRegistration)
class PledgeRegistrationAdmin(ModelAdmin):
    list_display = (
        "certificate_badge",
        "full_name",
        "mobile_number",
        "email",
        "organization",
        "gender",
        "city",
        "cert_status_badge",
        "email_sent_badge",
        "whatsapp_sent_badge",
        "created_at",
    )
    search_fields = (
        "certificate_id", "full_name", "mobile_number", "email",
        "organization__name", "city",
    )
    list_filter = (
        "organization",
        "gender",
        "otp_verified",
        "email_sent",
        "whatsapp_sent",
        ("created_at", admin.DateFieldListFilter),
    )
    date_hierarchy = "created_at"
    ordering = ("-created_at",)
    list_per_page = 25
    actions = ["export_excel_action", "mark_email_sent", "mark_whatsapp_sent"]

    fieldsets = (
        ("Personal Information", {
            "fields": (
                "full_name",
                ("mobile_number", "email"),
                ("date_of_birth", "age", "gender"),
                ("city", "pincode"),
            )
        }),
        ("Organization", {
            "fields": ("organization",)
        }),
        ("Status & Verification", {
            "fields": (
                "otp_verified",
                "consent_accepted",
                "is_approved",
                ("email_sent", "whatsapp_sent"),
            )
        }),
        ("Certificate", {
            "fields": (
                "certificate_id",
                "certificate_image",
                "certificate_pdf",
                "generated_at",
            )
        }),
        ("Analytics", {
            "fields": (
                "download_count",
                ("share_ig_count", "share_wa_count", "share_fb_count"),
                ("ip_address", "browser_info"),
            ),
            "classes": ("collapse",),
        }),
    )
    readonly_fields = ("age", "certificate_id", "generated_at")

    # ── Display Methods ────────────────────────────────────────────

    def certificate_badge(self, obj):
        if obj.certificate_id:
            return format_html(
                '<span style="font-family:monospace;font-size:0.75rem;background:#F0FDF4;'
                'color:#16A34A;padding:3px 8px;border-radius:6px;font-weight:700;">{}</span>',
                obj.certificate_id
            )
        return format_html('<span style="color:#9CA3AF;font-size:0.75rem;">—</span>')
    certificate_badge.short_description = "Cert ID"

    def cert_status_badge(self, obj):
        if obj.certificate_id:
            return format_html(
                '<span style="background:#DCFCE7;color:#16A34A;padding:2px 8px;'
                'border-radius:999px;font-size:0.7rem;font-weight:700;">✓ GENERATED</span>'
            )
        return format_html(
            '<span style="background:#FEF3C7;color:#D97706;padding:2px 8px;'
            'border-radius:999px;font-size:0.7rem;font-weight:700;">PENDING</span>'
        )
    cert_status_badge.short_description = "Certificate"

    def email_sent_badge(self, obj):
        if obj.email_sent:
            return format_html('<span style="color:#16A34A;font-size:1rem;" title="Email Sent">✉️</span>')
        return format_html('<span style="color:#D1D5DB;font-size:1rem;" title="Email Not Sent">✉️</span>')
    email_sent_badge.short_description = "Email"

    def whatsapp_sent_badge(self, obj):
        if obj.whatsapp_sent:
            return format_html('<span style="color:#16A34A;font-size:1rem;" title="WhatsApp Sent">📱</span>')
        return format_html('<span style="color:#D1D5DB;font-size:1rem;" title="WhatsApp Not Sent">📱</span>')
    whatsapp_sent_badge.short_description = "WA"

    # ── Actions ───────────────────────────────────────────────────

    @action(description=_("📊 Export selected to Excel (.xlsx)"))
    def export_excel_action(self, request, queryset):
        return _export_pledges_excel(queryset)

    @action(description=_("✉️ Mark selected as Email Sent"))
    def mark_email_sent(self, request, queryset):
        count = queryset.update(email_sent=True)
        messages.success(request, f"{count} registrations marked as email sent.")

    @action(description=_("📱 Mark selected as WhatsApp Sent"))
    def mark_whatsapp_sent(self, request, queryset):
        count = queryset.update(whatsapp_sent=True)
        messages.success(request, f"{count} registrations marked as WhatsApp sent.")


def _export_pledges_excel(queryset):
    """Generate a premium Excel file for pledge registrations."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pledges"
    ws.freeze_panes = "A2"

    header_fill = PatternFill(start_color="0B7A3B", end_color="0B7A3B", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = [
        "Certificate ID", "Full Name", "Mobile", "Email", "Gender",
        "Date of Birth", "Age", "City", "Pincode", "Organization",
        "OTP Verified", "Certificate Generated", "Email Sent", "WhatsApp Sent",
        "Download Count", "Registration Date",
    ]

    ws.append(headers)
    ws.row_dimensions[1].height = 28

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    for pledge in queryset:
        ws.append([
            pledge.certificate_id or "",
            pledge.full_name,
            pledge.mobile_number,
            pledge.email or "",
            pledge.get_gender_display() if pledge.gender else "",
            pledge.date_of_birth.strftime("%d-%m-%Y") if pledge.date_of_birth else "",
            pledge.age or "",
            pledge.city or "",
            pledge.pincode or "",
            pledge.organization.name if pledge.organization else "",
            "Yes" if pledge.otp_verified else "No",
            "Yes" if pledge.certificate_id else "No",
            "Yes" if pledge.email_sent else "No",
            "Yes" if pledge.whatsapp_sent else "No",
            pledge.download_count,
            pledge.created_at.strftime("%Y-%m-%d %H:%M") if pledge.created_at else "",
        ])

    # Auto width
    for col_idx, col in enumerate(ws.columns, 1):
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 40)

    today = datetime.date.today().strftime("%Y-%m-%d")
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="pledges_{today}.xlsx"'
    wb.save(response)
    return response
